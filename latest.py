# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'initial.ui'
#
# Created by: PyQt5 UI code generator 5.15.4
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox

class Ui_VAAHAN(object):
    def setupUi(self, VAAHAN):
        VAAHAN.setObjectName("VAAHAN")
        VAAHAN.resize(592, 541)
        font = QtGui.QFont()
        font.setFamily("Rockwell")
        font.setBold(True)
        font.setWeight(75)
        VAAHAN.setFont(font)
        VAAHAN.setAutoFillBackground(False)
        self.centralwidget = QtWidgets.QWidget(VAAHAN)
        self.centralwidget.setObjectName("centralwidget")
        self.textEdit = QtWidgets.QTextEdit(self.centralwidget)
        self.textEdit.setGeometry(QtCore.QRect(180, 310, 201, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.textEdit.setFont(font)
        self.textEdit.setObjectName("textEdit")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(120, 270, 321, 31))
        font = QtGui.QFont()
        font.setFamily("Segoe UI Semilight")
        font.setPointSize(14)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(210, 360, 141, 31))
        font = QtGui.QFont()
        font.setFamily("OCR A Std")
        font.setPointSize(12)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(170, 160, 231, 51))
        font = QtGui.QFont()
        font.setPointSize(40)
        self.label_3.setFont(font)
        self.label_3.setAutoFillBackground(True)
        self.label_3.setObjectName("label_3")
        self.label_3.setStyleSheet("background-color: cyan")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(0, 0, 601, 191))
        self.label_2.setText("")
        self.label_2.setPixmap(QtGui.QPixmap("header_700x290.jpg"))
        self.label_2.setObjectName("label_2")
        self.label_2.raise_()
        self.textEdit.raise_()
        self.label.raise_()
        self.pushButton.raise_()
        self.label_3.raise_()
        VAAHAN.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(VAAHAN)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 592, 21))
        self.menubar.setObjectName("menubar")
        self.menuStart_Detection = QtWidgets.QMenu(self.menubar)
        self.menuStart_Detection.setObjectName("menuStart_Detection")
        self.menuGet_Details = QtWidgets.QMenu(self.menubar)
        self.menuGet_Details.setObjectName("menuGet_Details")
        VAAHAN.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(VAAHAN)
        self.statusbar.setObjectName("statusbar")
        VAAHAN.setStatusBar(self.statusbar)
        self.menubar.addAction(self.menuStart_Detection.menuAction())
        self.menubar.addAction(self.menuGet_Details.menuAction())
        self.pushButton.clicked.connect(self.take_input)
        self.retranslateUi(VAAHAN)
        QtCore.QMetaObject.connectSlotsByName(VAAHAN)

    def retranslateUi(self, VAAHAN):
        _translate = QtCore.QCoreApplication.translate
        VAAHAN.setWindowTitle(_translate("VAAHAN", "MainWindow"))
        self.label.setText(_translate("VAAHAN", "Enter the Vehicle Registration Plate"))
        self.pushButton.setText(_translate("VAAHAN", "Get Details"))
        self.label_3.setText(_translate("VAAHAN", "VAAHAN"))
        self.menuStart_Detection.setTitle(_translate("VAAHAN", "Start Detection"))
        self.menuGet_Details.setTitle(_translate("VAAHAN", "Get Details"))

    def take_input(self, VAAHAN):
        mytext = self.textEdit.toPlainText()
        import openpyxl

        path = "output.xlsx"

        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active

        row = sheet_obj.max_row
        column = sheet_obj.max_column

        all_plate=[]
        timing=[]
        state=[]
        dating=[]
        auths=[]
        for i in range(1, row + 1):
                plate = sheet_obj.cell(row = i, column = 2)
                date=sheet_obj.cell(row=i,column=3)
                st=sheet_obj.cell(row=i,column=1)
                time=sheet_obj.cell(row=i,column=4)

                auth=sheet_obj.cell(row=i,column=5)

                if i>1:
                    all_plate.append(plate.value)
                    timing.append(time.value)
                    state.append(st.value)
                    dating.append(date.value)
                    auths.append(auth.value)



        # print(auths)

        import numpy as np
        timmee = np.array(all_plate)

        pl = mytext
        newms=''
        if pl in all_plate:
            print('')
            index = all_plate.index(pl)
            print('Number Plate is-',pl,"\n")
            st_idx= state[index]
            print(f'This vehicle is of state-{st_idx}',"\n")
            print(f'Timing stamps of this vehicle are-',"\n")
            ii = np.where(timmee == pl)[0]

            for i in ii:
                newms = newms + str(dating[i]) +'\t' + str(timing[i])+'\n'
                print(dating[i],timing[i])
            print('')
            print('AUTHENTICATED: ',auths[index])
        else:
            print('This vehicle has never passed from our system..')

        if pl in all_plate:
            newms = newms + '\n' + f'AUTHENTICATED : {auths[index]}'

            msg = f'Number Plate is - {pl}\n\nThis Vehicle is of State - {st_idx}\n\nTiming stamps of this vehicle are-\n\n{newms}'
            msgBox = QMessageBox()
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Information)
            msgBox.setText(msg)
            msgBox.setWindowTitle("Vehicle Details")
            msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            msgBox.exec_()
        else:
            msg = 'This vehicle has never passed from our system..'
            msgBox = QMessageBox()
            msgBox = QMessageBox()
            msgBox.setIcon(QMessageBox.Information)
            msgBox.setText(msg)
            msgBox.setWindowTitle("Vehicle Details")
            msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            msgBox.exec_()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    VAAHAN = QtWidgets.QMainWindow()
    ui = Ui_VAAHAN()
    ui.setupUi(VAAHAN)
    VAAHAN.show()
    sys.exit(app.exec_())
